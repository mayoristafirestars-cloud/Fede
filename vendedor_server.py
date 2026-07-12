"""
Vendedor virtual — atiende clientes del negocio por WhatsApp.

Lee la lista de precios de negocio/precios.xlsx y la info del negocio
de negocio/info.md. Puede mandar fotos de productos (negocio/fotos/).

POST /api/vendedor {"session_id": "...", "message": "..."}
  -> {"response": "...", "fotos": ["ruta1.jpg", ...]}

Correr: uvicorn vendedor_server:app --port 8003
"""
import json
import os
import re
import threading
import traceback
import unicodedata
from pathlib import Path

from anthropic import Anthropic
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel

from memoria import cargar_sesiones, guardar_sesiones, recortar_historial
from pagos import crear_link_pago, hay_pagos

load_dotenv()

# Modelo configurable: para recortar costo se puede usar
# EVA_MODEL=claude-haiku-4-5-20251001 (~4x más barato, respuestas más simples)
MODEL = os.getenv("EVA_MODEL", "claude-sonnet-4-6")
MAX_TOKENS = 1024
MAX_TURNS = 5

BASE_DIR = Path(__file__).parent
NEGOCIO_DIR = BASE_DIR / "negocio"
CSV_PATH = NEGOCIO_DIR / "inventario.csv"  # export de FactuPyme (prioridad)
EXCEL_PATH = NEGOCIO_DIR / "precios.xlsx"  # alternativa manual
INFO_PATH = NEGOCIO_DIR / "info.md"
FOTOS_DIR = NEGOCIO_DIR / "fotos"

client = Anthropic()
app = FastAPI(title="Vendedor virtual")
SESIONES_PATH = BASE_DIR / "sesiones_eva.json"
sessions: dict[str, list[dict]] = cargar_sesiones(SESIONES_PATH)

# Un lock por número: si el mismo cliente manda dos mensajes casi juntos,
# se procesan uno después del otro. Sin esto, dos hilos podrían mutar el
# mismo historial a la vez y romper el pareo tool_use/tool_result (la API
# lo rechaza con 400 y el cliente se queda sin respuesta).
_session_locks: dict[str, threading.Lock] = {}
_session_locks_guard = threading.Lock()


def _lock_sesion(session_id: str) -> threading.Lock:
    with _session_locks_guard:
        lk = _session_locks.get(session_id)
        if lk is None:
            lk = _session_locks[session_id] = threading.Lock()
        return lk


def normalizar(s: str) -> str:
    """minúsculas y sin acentos, para búsquedas tolerantes"""
    s = unicodedata.normalize("NFD", str(s).lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def cargar_productos() -> list[dict]:
    """Lee el Excel. Espera columnas: producto, precio y opcionalmente
    categoria, stock, descripcion, foto (nombres flexibles)."""
    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    encabezado = [normalizar(c or "") for c in next(filas)]

    productos = []
    for fila in filas:
        if not fila or all(v is None for v in fila):
            continue
        p = dict(zip(encabezado, fila))
        if not p.get("producto"):
            continue
        productos.append(
            {
                "producto": str(p.get("producto", "")).strip(),
                "categoria": str(p.get("categoria", "") or "").strip(),
                "precio": p.get("precio", ""),
                "stock": str(p.get("stock", "") or "").strip(),
                "descripcion": str(p.get("descripcion", "") or "").strip(),
                "foto": str(p.get("foto", "") or "").strip(),
            }
        )
    wb.close()
    return productos


def parsear_precio(valor) -> float:
    """'16,500.00' -> 16500.0 ; también acepta números directos."""
    if valor is None:
        return 0.0
    s = str(valor).strip().replace("$", "").replace(" ", "")
    if not s:
        return 0.0
    # formato FactuPyme: coma para miles, punto para decimales
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def cargar_inventario_csv() -> list[dict]:
    """Lee el export de FactuPyme (CSV separado por ';', encoding latin-1).
    Solo carga datos PÚBLICOS: nunca el precio de costo ni utilidades."""
    import csv

    productos = []
    with open(CSV_PATH, encoding="latin-1", newline="") as f:
        for fila in csv.DictReader(f, delimiter=";"):
            campos = {normalizar(k): (v or "").strip() for k, v in fila.items() if k}
            nombre = campos.get("descripcion", "").strip()
            if not nombre:
                continue
            precio = parsear_precio(campos.get("precio venta"))
            if precio <= 0:
                continue  # sin precio de venta -> no se ofrece
            # Lista 1 (consumidor final) = Precio Venta.
            # Lista 2 (mayorista) = costo * (1 + Utilidad 2 / 100).
            costo = parsear_precio(campos.get("precio costo"))
            u2 = parsear_precio(campos.get("utilidad 2"))
            # FactuPyme calcula Lista 2 = costo * (1 + U2/100) incluso con U2=0
            # (confirmado contra el reporte real de Lista 2 del negocio).
            precio_mayorista = round(costo * (1 + u2 / 100), 2) if costo > 0 else None
            try:
                cantidad = int(float(campos.get("cantidad", "0") or 0))
            except ValueError:
                cantidad = 0
            marca = campos.get("marca", "")
            if marca.lower() == "sin especificar":
                marca = ""
            foto = campos.get("imagenes", "")
            if "product.png" in foto:  # placeholder de FactuPyme, no es foto real
                foto = ""
            codigo = campos.get("codigo", "").replace("Cod:", "").strip()
            rubro = campos.get("rubro", "").strip()
            subrubro = campos.get("subrubro", "").strip()
            # Grabable: productos Nexo, EXCEPTO electrónica (no se puede grabar
            # un drone, un cargador, auriculares, etc.). Se puede sobrescribir
            # producto por producto desde negocio/productos_extra.csv.
            proveedor = campos.get("provedor", "") or campos.get("proveedor", "")
            es_nexo = "nexo" in proveedor.lower()
            texto_prod = normalizar(f"{nombre} {rubro} {subrubro}")
            NO_GRABABLE = (
                "drone", "cargador", "cable", "auricular", "airpad", "airpod",
                "parlante", "speaker", "bateria", "pila ", "power bank", "powerbank",
                "radio", "receptor", "mouse", "aro de led", "aros de led",
                "luz de noche", "humificador", "humidificador", "aspiradora",
                "reloj", "calculadora", "linterna", "walkie", "microfono", "usb",
                "fuente ", "adaptador", "led para cel",
            )
            personalizable = es_nexo and not any(k in texto_prod for k in NO_GRABABLE)
            productos.append(
                {
                    "codigo": codigo,
                    "producto": nombre,
                    "categoria": " / ".join(x for x in (rubro, subrubro) if x),
                    "marca": marca,
                    "precio_lista1_consumidor_final": precio,
                    "precio_lista2_mayorista": precio_mayorista,
                    "stock": "si" if cantidad > 0 else "no",
                    "cantidad": cantidad,
                    "descripcion": "",
                    "foto": foto,
                    "personalizable": personalizable,
                    "unidades_por_bulto": None,
                }
            )
    _aplicar_overlay(productos)
    return productos


EXTRA_PATH = NEGOCIO_DIR / "productos_extra.csv"


def _aplicar_overlay(productos: list[dict]) -> None:
    """Aplica datos que NO vienen de FactuPyme, cargados a mano en
    negocio/productos_extra.csv. Columnas (por código de producto):
      codigo, unidades_por_bulto, grabable
    - unidades_por_bulto: cuántas unidades trae la caja/bulto (número)
    - grabable: si/no  (fuerza si un producto se graba o no, más allá del automático)
    Solo hace falta cargar los productos que quieras ajustar; el resto
    queda con los valores automáticos."""
    if not EXTRA_PATH.is_file():
        return
    import csv

    por_codigo = {p["codigo"]: p for p in productos if p["codigo"]}
    aplicados = 0
    try:
        with open(EXTRA_PATH, encoding="utf-8-sig", newline="") as f:
            for fila in csv.DictReader(f):
                campos = {
                    normalizar(k).replace("_", " "): (v or "").strip()
                    for k, v in fila.items() if k
                }
                cod = campos.get("codigo", "").replace("Cod:", "").strip()
                p = por_codigo.get(cod)
                if not p:
                    continue
                bulto = campos.get("unidades por bulto", "") or campos.get("bulto", "")
                if bulto.isdigit():
                    p["unidades_por_bulto"] = int(bulto)
                grab = campos.get("grabable", "").lower()
                if grab in ("si", "sí", "true", "1"):
                    p["personalizable"] = True
                elif grab in ("no", "false", "0"):
                    p["personalizable"] = False
                aplicados += 1
        if aplicados:
            print(f"[vendedor] {aplicados} productos ajustados desde productos_extra.csv")
    except Exception as e:
        print(f"[vendedor] No pude leer productos_extra.csv: {e}")


_csv_mtime: float | None = None


def _cargar_inventario() -> list[dict]:
    global _csv_mtime
    if CSV_PATH.is_file():
        _csv_mtime = CSV_PATH.stat().st_mtime
        productos = cargar_inventario_csv()
        print(f"[vendedor] {len(productos)} productos cargados de {CSV_PATH.name}")
        return productos
    productos = cargar_productos()
    print(f"[vendedor] {len(productos)} productos cargados de {EXCEL_PATH.name}")
    return productos


def _refrescar_si_cambio() -> None:
    """Si el CSV fue reemplazado (nueva exportación de FactuPyme),
    recarga el inventario al vuelo — sin reiniciar a Eva."""
    global PRODUCTOS
    try:
        if CSV_PATH.is_file() and CSV_PATH.stat().st_mtime != _csv_mtime:
            PRODUCTOS = _cargar_inventario()
            print("[vendedor] Inventario actualizado en caliente ✔")
    except Exception as e:
        print(f"[vendedor] No pude recargar el inventario: {e}")


PRODUCTOS = _cargar_inventario()


def buscar_producto(consulta: str) -> dict:
    """Busca productos por texto (nombre o categoría)."""
    _refrescar_si_cambio()
    q = normalizar(consulta)
    palabras = [w for w in q.split() if len(w) > 1]
    resultados = []
    for p in PRODUCTOS:
        texto = normalizar(
            f"{p.get('codigo', '')} {p['producto']} {p['categoria']} "
            f"{p.get('marca', '')} {p['descripcion']}"
        )
        if not palabras or all(w in texto for w in palabras) or q in texto:
            resultados.append(p)
    if not resultados and palabras:
        # segunda pasada: alcanza con que matchee alguna palabra
        for p in PRODUCTOS:
            texto = normalizar(f"{p['producto']} {p['categoria']} {p.get('marca', '')}")
            if any(w in texto for w in palabras):
                resultados.append(p)
    # Inteligencia de demanda: registrar qué buscó el cliente y si había stock.
    con_stock = any(p.get("cantidad", 0) > 0 for p in resultados)
    registrar_busqueda(consulta, len(resultados), con_stock)
    return {"encontrados": len(resultados), "productos": resultados[:12]}


def listar_categorias() -> dict:
    _refrescar_si_cambio()
    cats = sorted({p["categoria"] for p in PRODUCTOS if p["categoria"]})
    return {"categorias": cats, "total_productos": len(PRODUCTOS)}


TOOLS = [
    {
        "name": "buscar_producto",
        "description": "Busca productos en la lista de precios por nombre o categoría. Devuelve producto, precio, stock, descripción y si tiene foto disponible.",
        "input_schema": {
            "type": "object",
            "properties": {
                "consulta": {
                    "type": "string",
                    "description": "Qué busca el cliente, ej: 'coca cola', 'cerveza', 'fernet'",
                }
            },
            "required": ["consulta"],
        },
    },
    {
        "name": "listar_categorias",
        "description": "Lista las categorías de productos disponibles y el total de productos.",
        "input_schema": {"type": "object", "properties": {}},
    },
]

TOOL_FUNCTIONS = {
    "buscar_producto": lambda args: buscar_producto(args.get("consulta", "")),
    "listar_categorias": lambda args: listar_categorias(),
}


# URL pública de la tienda online (para que Eva la difunda de vez en cuando).
TIENDA_URL = os.getenv("TIENDA_URL", "coronelsur.com.ar")


def build_system_prompt() -> str:
    info = INFO_PATH.read_text(encoding="utf-8")
    return f"""# IDENTIDAD

Sos **Eva**, la asistente de ventas de Dist Coronel Sur por WhatsApp. Atendés clientes las 24hs: respondés precios, stock, cómo comprar, y ayudás a armar pedidos. Te presentás como Eva.

# TONO

- Español argentino, tratás de "vos", amable y ágil como un buen vendedor.
- Respuestas CORTAS (es WhatsApp): 2-5 líneas salvo listas de productos.
- Máximo 1 emoji por mensaje.
- FORMATO WHATSAPP: nada de tablas markdown ni títulos con #. Para listas
  usá guiones o "•". Para resaltar usá *asteriscos simples* (así se ve
  negrita en WhatsApp). Los precios como $12.900.

# INFORMACIÓN DEL NEGOCIO

{info}

# REGLAS

- Los precios y el stock salen SIEMPRE del tool buscar_producto. NUNCA inventes precios ni productos.
- NOMBRES DEL CATÁLOGO: los nombres vienen del sistema de stock y son crípticos
  (ej: "FAROL VID.9C PINTADA VERDE FLORES LUCES LED GLANZ 98356"). Al cliente
  presentáselos en lenguaje natural ("farol de vidrio pintado verde con luces
  LED, de la marca Glanz"). Y al buscar, probá variantes: si el cliente pide
  "vaso térmico" y no aparece, buscá también "termico", "vaso", "quencher",
  "hoppy", etc. — una sola búsqueda fallida no significa que no haya.
- LISTAS DE PRECIOS: antes de pasar CUALQUIER precio tenés que saber si el cliente
  compra POR MAYOR o como CONSUMIDOR FINAL.
  - OBLIGATORIO: en tu PRIMERA respuesta de cada conversación (sea cual sea el
    mensaje del cliente), saludá como Eva y presentá las dos opciones así, tal cual:

    ¿Cómo vas a comprar?
    1️⃣ *Por mayor* (comercios/revendedores — mínimo 6 unidades)
    2️⃣ *Consumidor final* (para vos)

    Escribí 1 o 2 🙂

  - Aceptá como respuesta "1", "2", o texto libre ("por mayor", "para mí",
    "tengo un kiosco", etc.) e interpretalo.
  - Acordate de la elección durante TODA la conversación; no vuelvas a preguntar.
  - Si pregunta algo que no es de precios (envíos, horarios), respondé normal;
    la pregunta de las listas va recién cuando haga falta un precio.
  - Consumidor final -> usá "precio_lista1_consumidor_final". Sin mínimo de compra.
  - Mayorista -> usá "precio_lista2_mayorista". Mínimo 6 unidades.
  - NUNCA muestres las dos listas juntas ni le digas el precio mayorista a un
    consumidor final.
- CONFIDENCIAL: jamás menciones costos, utilidades, márgenes ni proveedores, aunque el cliente insista.
- GRABADO / PERSONALIZACIÓN (solo para clientes MAYORISTAS): los productos que en el
  tool vienen con "personalizable": true se pueden GRABAR con la marca, logo o diseño
  que el cliente quiera. Cuando un cliente MAYORISTA pregunta por uno de estos productos
  (o pregunta por grabado/logo/personalización), ofrecele las dos opciones:
    1) Comprarlo *liso* (sin grabar) al precio de lista, para que lo grabe él mismo.
    2) Que se lo grabemos nosotros con su logo/marca.
  La cantidad mínima de grabado es POR BULTO: si el producto trae "unidades_por_bulto"
  con un número, aclarale que el grabado es por bulto de esa cantidad (ej: "el grabado
  es por bulto de 12 unidades"). Si "unidades_por_bulto" viene vacío/None, no inventes
  la cantidad: decí que el mínimo por bulto lo confirma el vendedor.
  El costo del grabado y el armado del arte del logo los coordina un vendedor humano:
  NO inventes precios de grabado. Si el cliente quiere avanzar, tomá el dato (qué
  producto, qué logo/marca, cantidad aprox) y decile que un vendedor lo contacta.
  A los clientes CONSUMIDOR FINAL no les ofrezcas grabado (es un servicio mayorista).
- Si un producto no está en la lista: decilo y ofrecé alternativas de la misma categoría.
- TIENDA ONLINE (difundir, pero con moderación): tenemos catálogo web en {TIENDA_URL},
  con fotos y precios, y un carrito que termina en este mismo WhatsApp. CADA TANTO invitá
  al cliente a conocerla — pero como MUCHO UNA vez por conversación, y solo en un momento
  natural: cuando pregunta "¿qué tenés?" o por variedad, cuando hay más productos de los
  que entran en el chat, o al cerrar/despedirte. Ejemplo: "Si querés ver TODO el catálogo
  con fotos, entrá a {TIENDA_URL} 🛍️". NO la menciones en cada mensaje ni la repitas si ya
  la nombraste antes en esta charla. Y si el cliente CLARAMENTE ya viene de la web (te llega
  un pedido ya armado tipo carrito, o menciona la página), NO hace falta invitarlo: ya la
  conoce, seguí con su pedido.
- PEDIDO CONFIRMADO: cuando el cliente confirma su pedido, respondele el resumen
  (productos, cantidades, precios, total, y que Malcom lo va a contactar) y
  ADEMÁS agregá al final un bloque exacto así (no lo ve el cliente, va directo
  al vendedor humano):
  PEDIDO_CONFIRMADO:
  Cliente: (mayorista o consumidor final)
  - [código] producto x cantidad = subtotal
  TOTAL: $...
- FORMATO OBLIGATORIO PARA MOSTRAR PRODUCTOS: cuando muestres productos, NUNCA
  hagas una lista compacta de una línea por producto. SIEMPRE usá el formato
  detallado: cada producto en su bloque (nombre + código + precio + stock) y
  JUSTO DEBAJO una línea con el CÓDIGO del producto para su foto.
  Mostrá como máximo 3 productos así por mensaje (los 3 más relevantes). Si hay más,
  al final decí "tengo más opciones, ¿querés que te muestre otras?".

  MUY IMPORTANTE: la línea de foto se escribe con el CÓDIGO del producto (NO la URL).
  El sistema busca solo la foto correcta de ESE código, así la imagen SIEMPRE coincide
  con el producto. Escribila exactamente así (el cliente no ve el texto "FOTO:"):
  FOTO: <código del producto, ej: 39200>

  EJEMPLO EXACTO de una respuesta correcta mostrando 3 termos:

  Acá te muestro 3 termos 😊

  *Termo acero 1Lt media manija* (cód. 39200)
  • Precio: *$16.999* · ✅ hay stock
  FOTO: 39200

  *Kit Termo + Mate 1Lt* (cód. 37694)
  • Precio: *$54.000* · ✅ hay stock
  FOTO: 37694

  *Kit Termo Tereré 2.5Lt* (cód. 37712)
  • Precio: *$80.352* · ✅ hay stock
  FOTO: 37712

  ¿Te interesa alguno? También tengo repuestos y tapones si necesitás 🙂

- Poné la línea "FOTO: <código>" justo debajo de CADA producto que tenga foto. El
  código es el mismo que mostrás en "(cód. XXXXX)". Si el producto no tiene foto,
  no pongas la línea. Nunca inventes un código.
- Las ALTERNATIVAS o SUGERENCIAS van AL FINAL, después de los productos con sus
  fotos, y SIN foto (salvo que el cliente pida ver esas también).
- Cuando el cliente quiera CERRAR un pedido: resumí el pedido (productos, cantidades, total estimado) y decile que un vendedor humano lo contacta para confirmar stock, total final y entrega.
- Ante reclamos, cuentas corrientes o cosas fuera de tu alcance: derivá a humano.
- No des información que no esté en este documento o en la lista de precios.
"""


SYSTEM_PROMPT = build_system_prompt()


class VendedorRequest(BaseModel):
    session_id: str
    message: str
    telefono: str = ""  # número real del contacto (el bridge lo resuelve)


class AudioRequest(BaseModel):
    session_id: str
    audio_b64: str
    mimetype: str = "audio/ogg"
    telefono: str = ""


def _resolver_foto(nombre: str) -> str:
    """Resuelve una referencia de foto a su URL/ruta real.
    Acepta: un CÓDIGO de producto (busca su foto en el inventario, así la
    imagen SIEMPRE coincide con el producto), una URL http, o un archivo local."""
    nombre = nombre.strip()
    if not nombre:
        return ""
    # 1) Código de producto -> su foto del inventario (lo más confiable)
    codigo = nombre.replace("cód.", "").replace("cod.", "").replace("[", "").replace("]", "").strip()
    for p in PRODUCTOS:
        if p.get("codigo") and p["codigo"] == codigo:
            foto = p.get("foto", "")
            return foto if foto.startswith("http") else ""
    # 2) URL directa
    if nombre.startswith("http"):
        return nombre
    # 3) Archivo local en negocio/fotos/
    if (FOTOS_DIR / nombre).is_file():
        return str(FOTOS_DIR / nombre)
    return ""


def construir_secuencia(texto: str) -> tuple[str, list[str], list[dict]]:
    """Convierte la respuesta en una SECUENCIA ordenada de partes.

    CLAVE: cada foto se manda CON el texto de su producto como PIE (caption),
    en un solo mensaje. Así la descripción y la foto van pegadas y NUNCA se
    desordenan (en WhatsApp oficial, mensajes separados enviados rápido pueden
    llegar en distinto orden). El texto que quede sin foto (intro, alternativas)
    se manda como mensaje de texto normal.

    Devuelve:
      - texto_plano: todo el texto (para log/fallback)
      - fotos: lista de todas las fotos (fallback)
      - secuencia: partes en orden. Cada parte es:
          {"tipo": "texto", "contenido": "..."}                 -> mensaje de texto
          {"tipo": "foto", "url": "...", "caption": "..."}       -> imagen con pie
    """
    secuencia: list[dict] = []
    fotos: list[str] = []
    buffer: list[str] = []

    def flush_texto():
        contenido = "\n".join(buffer).strip()
        buffer.clear()
        if contenido:
            secuencia.append({"tipo": "texto", "contenido": contenido})

    for linea in texto.splitlines():
        m = re.match(r"^\s*FOTOS?\s*:\s*(.+)$", linea, flags=re.IGNORECASE)
        if m:
            resueltas = [_resolver_foto(n) for n in m.group(1).split(",")]
            resueltas = [r for r in resueltas if r]
            if resueltas:
                # El texto acumulado (la descripción de ESTE producto) se usa como
                # pie de la primera foto -> descripción y foto van juntas.
                caption = "\n".join(buffer).strip()
                buffer.clear()
                for i, url in enumerate(resueltas):
                    secuencia.append({
                        "tipo": "foto",
                        "url": url,
                        "caption": caption if i == 0 else "",
                    })
                    fotos.append(url)
        else:
            buffer.append(linea)
    flush_texto()

    texto_plano_parts = []
    for p in secuencia:
        if p["tipo"] == "texto":
            texto_plano_parts.append(p["contenido"])
        elif p.get("caption"):
            texto_plano_parts.append(p["caption"])
    texto_plano = "\n".join(texto_plano_parts).strip()
    return texto_plano, fotos, secuencia


# Sistema de gestión Coronel Sur (opcional): si estas variables están en
# el .env, cada pedido confirmado se crea como presupuesto en el sistema.
CORONEL_URL = os.getenv("CORONEL_URL", "").rstrip("/")
AGENTE_TOKEN = os.getenv("AGENTE_TOKEN", "eva-coronel-2026")


def _num_ar(s: str) -> float:
    """'157.080' -> 157080 ; '13.006,50' -> 13006.5 ; '6' -> 6"""
    s = s.strip().replace("$", "").replace(" ", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") == 1 and len(s.split(".")[1]) == 3:
        s = s.replace(".", "")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _precio_real(codigo: str, tipo_cliente: str) -> float | None:
    """Precio VERDADERO de un producto según la lista del cliente, tomado del
    inventario (NO de lo que escribió el modelo). Devuelve None si no se ubica
    el producto por código."""
    if not codigo:
        return None
    codigo = codigo.strip()
    es_mayor = "mayor" in (tipo_cliente or "").lower()
    for p in PRODUCTOS:
        if p.get("codigo") == codigo:
            if es_mayor and p.get("precio_lista2_mayorista"):
                return float(p["precio_lista2_mayorista"])
            return float(p.get("precio_lista1_consumidor_final") or 0) or None
    return None


def parsear_pedido(pedido: str) -> dict:
    """Convierte el bloque PEDIDO_CONFIRMADO en datos estructurados.

    IMPORTANTE: el precio de cada ítem se RECALCULA contra el inventario real
    (por código + lista del cliente), NO se confía en lo que escribió el modelo.
    Así lo que se factura/cobra siempre coincide con la lista de precios."""
    tipo = ""
    m = re.search(r"Cliente\s*:\s*(.+)", pedido, flags=re.IGNORECASE)
    if m:
        tipo = m.group(1).strip()
    items = []
    for linea in pedido.splitlines():
        m = re.match(
            r"^\s*[-•]\s*(?:\[(\w+)\])?\s*(.+?)\s*[xX×]\s*([\d.,]+)\s*=\s*\$?\s*([\d.,]+)",
            linea,
        )
        if m:
            codigo, desc, cant, sub = m.groups()
            cantidad = _num_ar(cant) or 1
            subtotal_texto = _num_ar(sub)
            # Precio real de la base; si no se ubica el producto, cae al del texto.
            precio_real = _precio_real(codigo, tipo)
            if precio_real is not None:
                precio_unitario = precio_real
            else:
                precio_unitario = round(subtotal_texto / cantidad, 2) if cantidad else subtotal_texto
                print(f"[pedido] ⚠️ Sin código en inventario ({desc.strip()[:30]}), uso precio del texto")
            items.append({
                "codigo": codigo or None,
                "descripcion": desc.strip(),
                "cantidad": cantidad,
                "precio_unitario": round(precio_unitario, 2),
                "subtotal": round(precio_unitario * cantidad, 2),
            })
    # El total SIEMPRE se recalcula sumando los ítems ya validados
    total = round(sum(i["subtotal"] for i in items), 2)
    return {"tipo_cliente": tipo, "items": items, "total": total}


def registrar_busqueda(consulta: str, resultados: int, con_stock: bool) -> None:
    """Registra en el sistema qué producto buscó un cliente (inteligencia de
    demanda). En segundo plano; nunca bloquea ni rompe la atención."""
    if not CORONEL_URL:
        return
    termino = (consulta or "").strip()
    if len(termino) < 2:
        return

    def _enviar():
        try:
            import httpx

            httpx.post(
                f"{CORONEL_URL}/api/agente/busqueda",
                json={
                    "termino": termino,
                    "resultados": resultados,
                    "con_stock": 1 if con_stock else 0,
                },
                headers={"X-Token": AGENTE_TOKEN},
                timeout=5,
            )
        except Exception:
            pass

    threading.Thread(target=_enviar, daemon=True).start()


def notificar_coronel(pedido_texto: str, session_id: str, telefono_real: str = "") -> str:
    """Crea el presupuesto en el sistema Coronel Sur (si está configurado).
    Devuelve el número de pedido (ej: EVA-00007) o '' si no se pudo crear."""
    if not CORONEL_URL:
        return ""
    try:
        import httpx

        datos = parsear_pedido(pedido_texto)
        if not datos["items"]:
            print("[pedido] No pude parsear items; no se envió al sistema.")
            return ""
        # Preferimos el número real del contacto; si no vino, los dígitos del ID.
        telefono = re.sub(r"[^0-9]", "", telefono_real) or re.sub(r"[^0-9]", "", session_id)
        r = httpx.post(
            f"{CORONEL_URL}/api/agente/pedido",
            json={"telefono": telefono, "texto_original": pedido_texto, **datos},
            headers={"X-Token": AGENTE_TOKEN},
            timeout=5,
        )
        if r.status_code == 200:
            numero = r.json().get("numero", "")
            print(f"[pedido] Presupuesto creado en el sistema: {numero}")
            return numero
        print(f"[pedido] El sistema respondió {r.status_code}: {r.text[:120]}")
    except Exception as e:
        print(f"[pedido] No se pudo notificar al sistema: {e}")
    return ""


def registrar_conversacion(session_id: str, telefono: str, mensaje: str,
                           respuesta: str, es_audio: bool = False) -> None:
    """Guarda el intercambio en el sistema (en segundo plano, no bloquea)."""
    if not CORONEL_URL:
        return

    def _enviar():
        try:
            import httpx

            httpx.post(
                f"{CORONEL_URL}/api/agente/conversacion",
                json={
                    "session": session_id,
                    "telefono": telefono,
                    "mensaje": mensaje,
                    "respuesta": respuesta,
                    "es_audio": 1 if es_audio else 0,
                },
                headers={"X-Token": AGENTE_TOKEN},
                timeout=5,
            )
        except Exception:
            pass  # el log nunca debe romper la atención al cliente

    threading.Thread(target=_enviar, daemon=True).start()


def extraer_pedido(texto: str) -> tuple[str, str]:
    """Separa el bloque 'PEDIDO_CONFIRMADO:' (para el vendedor humano)
    del texto visible para el cliente."""
    m = re.search(r"^\s*PEDIDO_CONFIRMADO\s*:?\s*$", texto, flags=re.IGNORECASE | re.MULTILINE)
    if not m:
        return texto, ""
    visible = texto[: m.start()].strip()
    pedido = texto[m.end():].strip()
    return visible, pedido


# El system prompt (personalidad + info del negocio) se cachea en la API:
# se cobra completo una vez cada 5 minutos y las reutilizaciones cuestan
# ~10% — con clientes activos baja el costo a una fracción.
SYSTEM_CACHEADO = [
    {"type": "text", "text": SYSTEM_PROMPT, "cache_control": {"type": "ephemeral"}}
]


def chat(historial: list[dict], mensaje: str) -> str:
    recortar_historial(historial, max_mensajes=30)
    historial.append({"role": "user", "content": mensaje})
    for _ in range(MAX_TURNS):
        response = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            system=SYSTEM_CACHEADO,
            tools=TOOLS,
            messages=historial,
        )
        historial.append({"role": "assistant", "content": response.content})
        if response.stop_reason != "tool_use":
            return "".join(b.text for b in response.content if b.type == "text")
        resultados = []
        for block in response.content:
            if block.type == "tool_use":
                fn = TOOL_FUNCTIONS.get(block.name)
                try:
                    out = fn(block.input) if fn else {"error": "tool desconocido"}
                except Exception as e:
                    out = {"error": str(e)}
                resultados.append(
                    {
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": json.dumps(out, ensure_ascii=False, default=str),
                    }
                )
        historial.append({"role": "user", "content": resultados})
    return "Dame un segundo que lo reviso y te confirmo 🙌"


def procesar_mensaje(session_id: str, texto: str, telefono: str = "",
                     es_audio: bool = False) -> dict:
    """Flujo completo: chat -> fotos -> pedido -> notificación y log al sistema.

    Todo el manejo de la sesión va bajo un lock POR NÚMERO (ver _lock_sesion):
    dos mensajes casi simultáneos del mismo cliente se procesan en orden y
    nunca corrompen el historial."""
    if texto.lower() == "reset":
        with _lock_sesion(session_id):
            sessions.pop(session_id, None)
        return {"response": "Conversación reiniciada.", "fotos": [], "pedido": ""}
    with _lock_sesion(session_id):
        historial = sessions.setdefault(session_id, [])
        try:
            respuesta = chat(historial, texto)
        except Exception as e:
            print("\n===== ERROR COMPLETO (sacale foto a esto) =====")
            traceback.print_exc()
            print("===============================================\n")
            raise HTTPException(status_code=500, detail=str(e)[:500])
        # Separar el bloque del pedido (va al vendedor humano, no al cliente)
        respuesta, pedido = extraer_pedido(respuesta)
        # Armar la secuencia ordenada: producto -> su foto -> producto -> su foto...
        limpio, fotos, secuencia = construir_secuencia(respuesta)
        if pedido:
            print(f"[pedido] Nuevo pedido confirmado:\n{pedido}\n")
            numero_pedido = notificar_coronel(pedido, session_id, telefono)
            # Link de pago Mercado Pago (si está configurado) -> última parte de texto.
            # La referencia es el número de pedido (EVA-xxxxx): así, cuando MP
            # avisa que se acreditó el pago, el sistema sabe qué pedido cerrar.
            if hay_pagos():
                datos = parsear_pedido(pedido)
                link = crear_link_pago(
                    datos["items"], referencia=numero_pedido or telefono or session_id)
                if link:
                    pago = (
                        f"💳 *Pagá online al instante acá:*\n{link}\n"
                        "_La mercadería queda reservada cuando se acredita el pago._"
                    )
                    secuencia.append({"tipo": "texto", "contenido": pago})
                    limpio += "\n\n" + pago
        registrar_conversacion(session_id, telefono, texto, limpio, es_audio)
        guardar_sesiones(sessions, SESIONES_PATH)
    return {
        "response": limpio or "🙌",
        "fotos": fotos,
        "secuencia": secuencia,
        "pedido": pedido,
    }


@app.post("/api/vendedor")
def api_vendedor(req: VendedorRequest):
    texto = req.message.strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Mensaje vacío.")
    return procesar_mensaje(req.session_id, texto, req.telefono)


@app.post("/api/vendedor/audio")
def api_vendedor_audio(req: AudioRequest):
    """Audio de WhatsApp: se transcribe y sigue el flujo normal."""
    import base64

    from transcripcion import transcribir, sufijo_por_mime

    try:
        audio = base64.b64decode(req.audio_b64)
    except Exception:
        raise HTTPException(status_code=400, detail="Audio inválido (base64).")
    if not audio:
        raise HTTPException(status_code=400, detail="Audio vacío.")
    try:
        texto = transcribir(audio, suffix=sufijo_por_mime(req.mimetype))
    except Exception as e:
        print("\n===== ERROR DE TRANSCRIPCION =====")
        traceback.print_exc()
        print("==================================\n")
        raise HTTPException(status_code=500, detail=f"Transcripción: {str(e)[:300]}")

    if not texto:
        return {
            "response": "No llegué a entender el audio 😅 ¿Me lo repetís o me lo escribís?",
            "fotos": [], "pedido": "", "transcript": "",
        }
    print(f"[audio] Cliente dijo: {texto[:120]}")
    resultado = procesar_mensaje(req.session_id, texto, req.telefono, es_audio=True)
    resultado["transcript"] = texto
    return resultado


@app.get("/health")
def health():
    return {"ok": True, "productos": len(PRODUCTOS)}
